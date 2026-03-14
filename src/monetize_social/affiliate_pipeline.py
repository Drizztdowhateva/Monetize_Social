import csv
import ipaddress
import json
import socket
import shutil
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


@dataclass
class PipelinePaths:
    root: Path

    @property
    def templates(self) -> Path:
        return self.root / "data/templates"

    @property
    def exports(self) -> Path:
        return self.root / "data/exports"

    @property
    def snapshots(self) -> Path:
        return self.root / "data/snapshots"

    @property
    def master(self) -> Path:
        return self.templates / "affiliate_master_fillable.csv"

    @property
    def outreach(self) -> Path:
        return self.templates / "outreach_crm_lite.csv"

    @property
    def taxes(self) -> Path:
        return self.templates / "taxes_tracker_llc_1099_sole_prop.csv"

    @property
    def priority_csv(self) -> Path:
        return self.exports / "affiliate_priority_apply_list.csv"

    @property
    def payout_csv(self) -> Path:
        return self.exports / "affiliate_high_payout_potential.csv"

    @property
    def dashboard_csv(self) -> Path:
        return self.exports / "monthly_dashboard_summary.csv"

    @property
    def validation_csv(self) -> Path:
        return self.exports / "validation_issues.csv"

    @property
    def workbook(self) -> Path:
        return self.exports / "affiliate_tracker_workbook.xlsx"

    @property
    def changelog(self) -> Path:
        return self.snapshots / "changelog.md"


priority_bias = {
    "ClickBank": 10,
    "ShareASale": 9,
    "Amazon Associates": 9,
    "FlexOffers": 8,
    "CJ Affiliate": 8,
    "Awin": 7,
    "Impact": 7,
    "PartnerStack": 7,
    "Google Workspace Affiliate": 6,
    "Rakuten Advertising": 5,
    "MaxBounty": 5,
    "Shopify Collabs": 8,
    "YouTube Shopping Affiliate": 7,
    "TikTok Shop Affiliate": 7,
}

difficulty_score = {
    "low": 4,
    "low-medium": 3,
    "medium": 2,
    "medium-high": 1,
    "high": 0,
}

payout_bias = {
    "PartnerStack": 10,
    "Impact": 9,
    "CJ Affiliate": 8,
    "Awin": 8,
    "Rakuten Advertising": 8,
    "Google Workspace Affiliate": 8,
    "MaxBounty": 9,
    "ClickBank": 8,
    "TikTok Shop Affiliate": 8,
    "YouTube Shopping Affiliate": 7,
    "Shopify Collabs": 7,
    "Amazon Associates": 7,
    "ShareASale": 7,
    "FlexOffers": 7,
}

priority_columns = [
    "Platform",
    "Category",
    "Apply/Signup URL",
    "Primary Benefits",
    "Approval Difficulty",
    "Geo Availability",
    "Best For",
    "Program Status",
    "Program Health Score",
    "My Website/Channel",
    "Application Submitted (Y/N)",
    "Application Date",
    "Approval Status (Pending/Approved/Rejected)",
    "Affiliate ID / Publisher ID",
    "Store ID / Associate Tag",
    "Next Action",
    "Next Follow-up Date",
]

payout_columns = [
    "Platform",
    "Category",
    "Official Program URL",
    "Commission Model",
    "Cookie Duration",
    "Typical Payout Timing",
    "Payment Methods",
    "Primary Benefits",
    "Best For",
    "Program Status",
    "Program Health Score",
    "My Primary Niche",
    "Niche Fit Score (1-10)",
    "Top Offers To Promote",
    "Primary Landing Page",
    "Internal Notes",
]


def normalize(text: str) -> str:
    return (text or "").strip().lower()


def parse_int(value: str, default: int = 0) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def is_yes(value: str) -> bool:
    return normalize(value) in {"y", "yes", "true", "1"}


def valid_url_format(url: str) -> bool:
    if not url:
        return False
    parsed = urlparse(url)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def score_priority(row: dict) -> int:
    name = row.get("Platform", "").strip()
    diff = normalize(row.get("Approval Difficulty", ""))
    status = normalize(row.get("Program Status", ""))

    score = priority_bias.get(name, 4)
    score += difficulty_score.get(diff, 1)

    if "active" in status:
        score += 2
    if "unclear" in status:
        score -= 3

    return score


def score_payout(row: dict) -> int:
    name = row.get("Platform", "").strip()
    model = normalize(row.get("Commission Model", ""))

    score = payout_bias.get(name, 5)
    if "recurring" in model:
        score += 2
    if "cpa" in model or "cpl" in model:
        score += 2
    if "% of sale" in model:
        score += 1
    if "unknown" in model:
        score -= 2

    return score


def compliance_gaps(row: dict) -> str:
    gaps = []
    if not is_yes(row.get("Tax Form Submitted (Y/N)", "")):
        gaps.append("tax_form")
    if not is_yes(row.get("Payment Profile Completed (Y/N)", "")):
        gaps.append("payment_profile")
    if not is_yes(row.get("Promo Assets Access (Y/N)", "")):
        gaps.append("promo_assets")
    if not is_yes(row.get("2FA Enabled (Y/N)", "")):
        gaps.append("2fa")
    if not row.get("Compliance Notes", "").strip():
        gaps.append("compliance_notes")
    return ",".join(gaps)


def health_score(row: dict) -> int:
    base = 50
    base += score_priority(row)
    base += score_payout(row)
    base += max(0, min(10, parse_int(row.get("Niche Fit Score (1-10)", 0)))) * 2

    if is_yes(row.get("Application Submitted (Y/N)", "")):
        base += 4

    approval = normalize(row.get("Approval Status (Pending/Approved/Rejected)", ""))
    if "approved" in approval:
        base += 12
    elif "pending" in approval:
        base += 3
    elif "rejected" in approval:
        base -= 12

    if compliance_gaps(row):
        base -= 8

    return max(0, min(100, base))


_INJECTION_CHARS = {"=", "+", "-", "@", "\t", "\r"}


def _sanitize_csv_cell(value: object) -> str:
    """Prefix formula-injection characters so spreadsheet apps don't execute them."""
    text = "" if value is None else str(value)
    if text and text[0] in _INJECTION_CHARS:
        return "'" + text
    return text


def _to_excel_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value)
    return str(value)


def write_csv(path: Path, rows: list[dict], columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: _sanitize_csv_cell(row.get(col, "")) for col in columns})


def add_data_validation(ws, header: list[str]) -> None:
    col = {name: idx + 1 for idx, name in enumerate(header)}
    max_row = max(1000, ws.max_row + 200)

    def apply_list(col_name: str, values: list[str]):
        if col_name not in col:
            return
        letter = get_column_letter(col[col_name])
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{max_row}")

    apply_list("Application Submitted (Y/N)", ["Y", "N"])
    apply_list("Promo Assets Access (Y/N)", ["Y", "N"])
    apply_list("Tax Form Submitted (Y/N)", ["Y", "N"])
    apply_list("Payment Profile Completed (Y/N)", ["Y", "N"])
    apply_list("2FA Enabled (Y/N)", ["Y", "N"])
    apply_list("Approval Status (Pending/Approved/Rejected)", ["Pending", "Approved", "Rejected"])
    apply_list("Program Status", ["Active", "Paused", "Unknown", "Unclear public affiliate signup"])
    apply_list("Approval Difficulty", ["Low", "Low-Medium", "Medium", "Medium-High", "High"])


def write_sheet(ws, rows: list[dict], columns: list[str]) -> None:
    columns = [c for c in columns if isinstance(c, str) and c]
    ws.append(columns)

    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(bold=True)

    for col_index, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    for row in rows:
        ws.append([_to_excel_cell(row.get(col_name, "")) for col_name in columns])

    ws.freeze_panes = "A2"
    add_data_validation(ws, columns)

    sample_max = min(ws.max_row, 201)  # header + up to 200 data rows
    for i, col_name in enumerate(columns, start=1):
        max_len = max(14, len(col_name))
        for row_cells in ws.iter_rows(min_row=2, max_row=sample_max, min_col=i, max_col=i):
            value = row_cells[0].value
            text = "" if value is None else str(value)
            max_len = max(max_len, min(60, len(text)))
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2


def _is_safe_url(url: str) -> bool:
    """Return False for URLs that resolve to private/loopback addresses (SSRF guard)."""
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    hostname = parsed.hostname or ""
    if not hostname:
        return False
    try:
        addr = ipaddress.ip_address(socket.gethostbyname(hostname))
        if addr.is_private or addr.is_loopback or addr.is_link_local or addr.is_reserved:
            return False
    except (OSError, ValueError):
        return False
    return True


def _check_one_url(idx: int, platform: str, field: str, url: str) -> dict | None:
    """HEAD-check a single URL; return an issue dict or None."""
    if not _is_safe_url(url):
        return {"row": idx, "platform": platform, "issue_type": "blocked_url",
                "detail": f"{field}: resolved to private/reserved address"}
    try:
        req = Request(url, headers={"User-Agent": "MonetizeSocial/1.0"}, method="HEAD")
        with urlopen(req, timeout=6) as response:  # noqa: S310  (scheme validated above)
            code = getattr(response, "status", 200)
        if code >= 400:
            return {"row": idx, "platform": platform, "issue_type": "broken_link",
                    "detail": f"{field}: HTTP {code}"}
    except Exception as exc:  # noqa: BLE001
        return {"row": idx, "platform": platform, "issue_type": "broken_link",
                "detail": f"{field}: {exc}"}
    return None


def validate_rows(rows: list[dict], check_live_links: bool = False) -> list[dict]:
    issues = []

    platform_counts = Counter((r.get("Platform", "").strip().lower() for r in rows if r.get("Platform", "").strip()))

    live_link_tasks: list[tuple[int, str, str, str]] = []

    for idx, row in enumerate(rows, start=2):
        platform = row.get("Platform", "").strip()
        if not platform:
            issues.append({"row": idx, "platform": "", "issue_type": "missing_platform", "detail": "Platform is empty"})
            continue

        if platform_counts[platform.lower()] > 1:
            issues.append(
                {
                    "row": idx,
                    "platform": platform,
                    "issue_type": "duplicate_platform",
                    "detail": "Duplicate platform name found",
                }
            )

        for field in ["Official Program URL", "Apply/Signup URL", "Support/Docs URL"]:
            url = row.get(field, "").strip()
            if url and not valid_url_format(url):
                issues.append(
                    {
                        "row": idx,
                        "platform": platform,
                        "issue_type": "invalid_url_format",
                        "detail": f"{field}: {url}",
                    }
                )

            if check_live_links and url and valid_url_format(url):
                live_link_tasks.append((idx, platform, field, url))

    if live_link_tasks:
        with ThreadPoolExecutor(max_workers=8) as pool:
            futures = {
                pool.submit(_check_one_url, idx, platform, field, url): None
                for idx, platform, field, url in live_link_tasks
            }
            for future in as_completed(futures):
                result = future.result()
                if result is not None:
                    issues.append(result)

    return issues


def monthly_dashboard_rows(rows: list[dict], issues: list[dict]) -> list[dict]:
    total = len(rows)
    submitted = sum(1 for r in rows if is_yes(r.get("Application Submitted (Y/N)", "")))
    approved = sum(
        1 for r in rows if "approved" in normalize(r.get("Approval Status (Pending/Approved/Rejected)", ""))
    )
    pending = sum(
        1 for r in rows if "pending" in normalize(r.get("Approval Status (Pending/Approved/Rejected)", ""))
    )
    rejected = sum(
        1 for r in rows if "rejected" in normalize(r.get("Approval Status (Pending/Approved/Rejected)", ""))
    )

    avg_health = round(sum(health_score(r) for r in rows) / total, 2) if total else 0
    gaps = sum(1 for r in rows if compliance_gaps(r))

    data = [
        {"Metric": "Total Programs", "Value": total},
        {"Metric": "Applications Submitted", "Value": submitted},
        {"Metric": "Approved", "Value": approved},
        {"Metric": "Pending", "Value": pending},
        {"Metric": "Rejected", "Value": rejected},
        {"Metric": "Average Health Score", "Value": avg_health},
        {"Metric": "Programs With Compliance Gaps", "Value": gaps},
        {"Metric": "Validation Issues", "Value": len(issues)},
    ]

    return data


def snapshot_master(paths: PipelinePaths) -> Path:
    paths.snapshots.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    out = paths.snapshots / f"affiliate_master_snapshot_{stamp}.csv"
    shutil.copy2(paths.master, out)
    return out


def append_changelog(paths: PipelinePaths, summary: dict) -> None:
    paths.snapshots.mkdir(parents=True, exist_ok=True)
    now = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%SZ")
    lines = [
        f"## {now}",
        f"- total_programs: {summary['total_programs']}",
        f"- avg_health_score: {summary['avg_health_score']}",
        f"- validation_issues: {summary['validation_issues']}",
        f"- snapshot: {summary['snapshot']}",
        "",
    ]
    if not paths.changelog.exists():
        paths.changelog.write_text("# Change Log\n\n", encoding="utf-8")
    with paths.changelog.open("a", encoding="utf-8") as fh:
        fh.write("\n" + "\n".join(lines))


def ensure_onboarding_packets(rows: list[dict], root: Path) -> None:
    packets = root / "docs/onboarding_packets"
    packets.mkdir(parents=True, exist_ok=True)
    packets_resolved = packets.resolve()

    for row in rows:
        platform = row.get("Platform", "").strip()
        if not platform:
            continue
        folder_name = "".join(c if c.isalnum() or c in {"-", "_"} else "_" for c in platform).strip("_")
        if not folder_name:
            continue
        folder = packets / folder_name.lower()
        # Path traversal guard: ensure resolved path stays inside packets dir
        if not folder.resolve().is_relative_to(packets_resolved):
            continue
        folder.mkdir(parents=True, exist_ok=True)

        checklist = folder / "checklist.md"
        if not checklist.exists():
            checklist.write_text(
                "\n".join(
                    [
                        f"# {platform} Onboarding Packet",
                        "",
                        "- [ ] Review terms and restrictions",
                        "- [ ] Submit application",
                        "- [ ] Record affiliate/publisher ID",
                        "- [ ] Configure tax and payment profile",
                        "- [ ] Add disclosure and compliance notes",
                        "- [ ] Build first deep link",
                        "- [ ] Set next follow-up date",
                    ]
                ),
                encoding="utf-8",
            )

        info = folder / "program_info.json"
        info.write_text(json.dumps(row, indent=2), encoding="utf-8")


def build_workbook(paths: PipelinePaths, rows: list[dict], issues: list[dict]) -> None:
    priority_rows = sorted(rows, key=score_priority, reverse=True)
    payout_rows = sorted(rows, key=score_payout, reverse=True)

    for row in rows:
        row["Program Health Score"] = health_score(row)
        row["Compliance Gaps"] = compliance_gaps(row)

    paths.exports.mkdir(parents=True, exist_ok=True)

    write_csv(paths.priority_csv, priority_rows, priority_columns)
    write_csv(paths.payout_csv, payout_rows, payout_columns)
    write_csv(paths.validation_csv, issues, ["row", "platform", "issue_type", "detail"])

    dashboard = monthly_dashboard_rows(rows, issues)
    write_csv(paths.dashboard_csv, dashboard, ["Metric", "Value"])

    wb = Workbook()
    wb.remove(wb.active)

    all_columns = [c for c in list(rows[0].keys()) if isinstance(c, str) and c] if rows else []

    write_sheet(wb.create_sheet("Master"), rows, all_columns)
    write_sheet(wb.create_sheet("Priority Apply"), priority_rows, priority_columns)
    write_sheet(wb.create_sheet("High Payout"), payout_rows, payout_columns)
    write_sheet(wb.create_sheet("Dashboard"), dashboard, ["Metric", "Value"])
    write_sheet(wb.create_sheet("Validation Issues"), issues, ["row", "platform", "issue_type", "detail"])

    applied = [r for r in rows if is_yes(r.get("Application Submitted (Y/N)", ""))]
    approved = [r for r in rows if "approved" in normalize(r.get("Approval Status (Pending/Approved/Rejected)", ""))]
    rejected = [r for r in rows if "rejected" in normalize(r.get("Approval Status (Pending/Approved/Rejected)", ""))]

    write_sheet(wb.create_sheet("Applied"), applied, all_columns)
    write_sheet(wb.create_sheet("Approved"), approved, all_columns)
    write_sheet(wb.create_sheet("Rejected"), rejected, all_columns)

    if paths.outreach.exists():
        with paths.outreach.open(newline="", encoding="utf-8") as f:
            outreach_rows = list(csv.DictReader(f))
        outreach_cols = list(outreach_rows[0].keys()) if outreach_rows else ["Brand", "Status"]
        write_sheet(wb.create_sheet("Outreach CRM"), outreach_rows, outreach_cols)

    if paths.taxes.exists():
        with paths.taxes.open(newline="", encoding="utf-8") as f:
            tax_rows = list(csv.DictReader(f))
        tax_cols = list(tax_rows[0].keys()) if tax_rows else ["Tax Year", "Platform"]
        write_sheet(wb.create_sheet("Taxes Tracker"), tax_rows, tax_cols)

    wb.save(paths.workbook)


def run_pipeline(root: Path, live_link_check: bool = False) -> dict:
    paths = PipelinePaths(root=root)
    paths.exports.mkdir(parents=True, exist_ok=True)

    with paths.master.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    if not rows:
        raise ValueError(f"No rows found in {paths.master}")

    snapshot_path = snapshot_master(paths)
    issues = validate_rows(rows, check_live_links=live_link_check)

    # Compute once per row; reused by build_workbook and dashboard summary
    for row in rows:
        row["Program Health Score"] = health_score(row)
        row["Compliance Gaps"] = compliance_gaps(row)

    build_workbook(paths, rows, issues)
    ensure_onboarding_packets(rows, paths.root)

    avg = round(sum(int(r.get("Program Health Score", 0)) for r in rows) / len(rows), 2)
    summary = {
        "total_programs": len(rows),
        "avg_health_score": avg,
        "validation_issues": len(issues),
        "snapshot": str(snapshot_path),
    }
    append_changelog(paths, summary)
    return summary


def add_workbook_validations(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        add_data_validation(ws, header)
    wb.save(path)
